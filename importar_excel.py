from openpyxl import load_workbook

from calculos import calcular_preco
from banco import salvar_produto


def importar_produtos_excel(caminho_arquivo):
    planilha = load_workbook(caminho_arquivo)

    aba = planilha.active

    for linha in aba.iter_rows(min_row=2, values_only=True):

        nome = linha[0]

        custo_ingredientes = float(linha[1])

        embalagem = float(linha[2])

        mao_obra = float(linha[3])

        margem_lucro = float(linha[4])

        estoque = int(linha[5])

        custo_total, lucro, preco_venda = calcular_preco(
            custo_ingredientes,
            embalagem,
            mao_obra,
            margem_lucro
        )

        salvar_produto(
            nome,
            custo_ingredientes,
            embalagem,
            mao_obra,
            margem_lucro,
            custo_total,
            lucro,
            preco_venda,
            estoque
        )